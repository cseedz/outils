#!/usr/bin/env python3
"""
casting_watcher.py
Watcher Casting Sauvage — Google Sheets → Supabase Storage → Supabase

Modes :
  python3 casting_watcher.py check <mcp_xlsx_file>
      Parse le XLSX, compare avec Supabase, imprime JSON des nouveaux candidats.
      Chaque entrée inclut un champ "extra" avec toutes les données du formulaire.

  python3 casting_watcher.py save <nom> <age> <role> <selectors> <photo_url> [extra_json]
      Ajoute un candidat dans Supabase (casting_talents, agence=Casting Sauvage).
      extra_json (optionnel) : JSON des données formulaire → stocké dans fiche_cs.
      ET l'ajoute dans les shortlists de personnages correspondant au rôle.

  python3 casting_watcher.py process-photo <mcp_photo_file> <nom>
      Extrait le base64 du fichier MCP, convertit si HEIC, uploade sur Cloudinary,
      imprime l'URL Cloudinary.

  python3 casting_watcher.py enrich <mcp_xlsx_file>
      Enrichit rétroactivement les candidats CS existants avec les données du formulaire
      (ajoute fiche_cs sur chaque talent qui n'en a pas encore).

  python3 casting_watcher.py sync-shortlists
      Rattache rétroactivement tous les candidats Casting Sauvage existants
      aux shortlists de personnages selon leur champ notes (rôle).
"""

import sys
import json
import re
import base64
import tempfile
import subprocess
import os
import requests
from datetime import datetime, timezone

# ── Config ────────────────────────────────────────────────────────────────────
SUPABASE_URL = 'https://dvivafrldxzhkactsvve.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImR2aXZhZnJsZHh6aGthY3RzdnZlIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzkyMzQ5MDgsImV4cCI6MjA5NDgxMDkwOH0.EgyxWDERi443hefaM0LxDYDhLWQYx31feKzQ1bQU5Kc'
CASTING_KEY  = 'casting_talents'
PERSO_KEY    = 'premier-trio_personnages'
STORAGE_KEY    = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImR2aXZhZnJsZHh6aGthY3RzdnZlIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3OTIzNDkwOCwiZXhwIjoyMDk0ODEwOTA4fQ.ZBboJewOeyBTfj4QG9i3frwUFZuoMLQMN1v9jO4x1i0'
STORAGE_BUCKET = 'photos'
SHEETS_FILE_ID    = '1mBjSLSUSBruYxcpPdTtapdjAME9vFl6besn950xEWCA'

# Mapping mot-clé (minuscules) → clé personnage dans premier-trio_personnages
_ROLE_CHAR_MAP = {
    'louis':    'LOUIS',
    'arnaud':   'ARNAUD',
    'thomas':   'THOMAS',
    'zack':     'ZACK',
    'charlotte':'CHARLOTTE',
    'sophia':   'SOPHIA',
    'raph':     'RAPH',
    'chloé':    'CHLOÉ',
    'chloe':    'CHLOÉ',
    'xavier':   'XAVIER',
    'émile':    'ÉMILE',
    'emile':    'ÉMILE',
}
# ─────────────────────────────────────────────────────────────────────────────

SB_HEADERS = {
    'apikey': SUPABASE_KEY,
    'Authorization': f'Bearer {SUPABASE_KEY}',
    'Content-Type': 'application/json',
}


def extract_xlsx_from_mcp(mcp_file: str) -> bytes:
    """Extrait le contenu XLSX base64 d'un fichier de résultat MCP Drive."""
    with open(mcp_file, 'rb') as f:
        f.read(12)  # {"content":"
        b64 = bytearray()
        prev = b''
        while True:
            chunk = f.read(65536)
            if not chunk:
                break
            combined = prev + chunk
            idx = combined.find(b'","id"')
            if idx != -1:
                b64.extend(combined[:idx])
                break
            b64.extend(prev)
            prev = chunk
    return base64.b64decode(b64.decode('ascii'))


def extract_img_from_mcp(mcp_file: str) -> bytes:
    """Extrait les bytes image base64 d'un fichier de résultat MCP Drive."""
    with open(mcp_file, 'rb') as f:
        f.read(12)
        b64 = bytearray()
        prev = b''
        while True:
            chunk = f.read(65536)
            if not chunk:
                break
            combined = prev + chunk
            idx = combined.find(b'","id"')
            if idx != -1:
                b64.extend(combined[:idx])
                break
            b64.extend(prev)
            prev = chunk
    return base64.b64decode(b64.decode('ascii'))


def is_green(cell):
    fill = cell.fill
    if not fill or fill.fill_type != 'solid':
        return False
    fc = fill.fgColor
    if fc.type == 'theme' and fc.theme == 7:
        return True
    if fc.type == 'rgb':
        try:
            return str(fc.rgb) == 'FFB6D7A8'
        except Exception:
            pass
    return False


def _parse_extra_from_response_row(rep, row) -> dict:
    """Extrait les données riches d'une ligne de la feuille Réponses (colonnes 4–32)."""
    def cell(col):
        v = rep.cell(row, col).value
        return str(v).strip() if v is not None else ''

    extra = {}
    mapping = [
        (4,  'ville'),
        (6,  'tel'),
        (7,  'email'),
        (8,  'parent_nom'),
        (9,  'parent_tel'),
        (10, 'parent_email'),
        (13, 'a_tourne'),
        (14, 'projets_tourne'),
        (15, 'uda'),
        (16, 'hockey'),
        (17, 'hockey_niveau'),
        (18, 'hockey_annees'),
        (19, 'hockey_equipe'),
        (20, 'hockey_position'),
        (21, 'danse'),
        (22, 'danse_ecole'),
        (23, 'danse_type'),
        (24, 'danse_annees'),
        (25, 'indispo_auditions'),
        (26, 'indispo_auditions_detail'),
        (27, 'indispo_tournage'),
        (28, 'indispo_tournage_detail'),
        (29, 'video_url'),
        (32, 'photo2_url_drive'),
    ]
    for col, key in mapping:
        v = cell(col)
        if v:
            extra[key] = v
    return extra


def parse_age(s):
    if not s:
        return None
    m = re.search(r'(\d+)', str(s))
    return int(m.group(1)) if m else None


def guess_genre(prenom):
    fem  = ['laura','marie','mariane','maëva','maya','marguerite','rosalie','emma',
            'rose','ann','alicia','juliette','julianne','romy','lauralou','berthelette',
            'alexia','clara','océanne','lauriane','maude','nadia','gabrielle','emy',
            'émie','amélie','élisabeth','léa','sarah','sofia','jade','yasmine','lily',
            'mia','charlotte','sophia','zoé','alice','camille','manon','lucie','inès']
    masc = ['samuel','thomas','charles','matys','matisse','milan','lou-félyx','zachary',
            'zackary','noah','alexis','lou','félix','olivier','maxime','antoine','william',
            'gabriel','nathan','ethan','jacob','liam','logan','ryan','xavier','victor',
            'adam','alexandre','nicolas','raphaël','édouard','arnaud','pierre','marc']
    n = prenom.lower().strip()
    if any(n.startswith(f) for f in fem):
        return 'femme'
    if any(n.startswith(m) for m in masc):
        return 'homme'
    return None


def load_existing_names() -> set:
    """Charge les noms déjà dans Casting Sauvage depuis Supabase."""
    r = requests.get(
        f'{SUPABASE_URL}/rest/v1/project_data',
        params={'key': f'eq.{CASTING_KEY}', 'select': 'data'},
        headers=SB_HEADERS, timeout=15,
    )
    r.raise_for_status()
    rows = r.json()
    if not rows:
        return set()
    talents = rows[0]['data'].get('talents', [])
    return {t['nom'].strip().lower() for t in talents if t.get('agence') == 'Casting Sauvage'}


def cmd_check(mcp_xlsx_file: str):
    """
    Parse le XLSX, trouve les nouveaux candidats verts, imprime JSON.
    Chaque entrée : {nom, age, role, selectors, file_id}
    """
    try:
        import openpyxl
    except ImportError:
        sys.exit('openpyxl requis : pip3 install openpyxl')

    xlsx_bytes = extract_xlsx_from_mcp(mcp_xlsx_file)
    xlsx_path  = '/tmp/_casting_watch_tmp.xlsx'
    with open(xlsx_path, 'wb') as f:
        f.write(xlsx_bytes)

    wb  = openpyxl.load_workbook(xlsx_path, data_only=True)
    sel = wb['SÉLECTIONS']
    rep = wb['Réponses au formulaire 1']

    # Index Réponses : (prenom.lower, nom.lower) → {photo1, extra}
    # Dernier envoi gagne (overwrite) — si quelqu'un a re-soumis, on prend le plus récent.
    resp_index = {}
    for row in range(2, rep.max_row + 1):
        prenom = rep[f'B{row}'].value
        nom    = rep[f'C{row}'].value
        if not prenom and not nom:
            continue
        key    = (str(prenom).strip().lower(), str(nom).strip().lower())
        photo1 = rep.cell(row=row, column=31).value
        resp_index[key] = {
            'photo1': str(photo1) if photo1 else None,
            'extra':  _parse_extra_from_response_row(rep, row),
        }

    existing = load_existing_names()

    new_candidates = []
    for row in range(2, sel.max_row + 1):
        prenom = sel[f'A{row}'].value
        nom    = sel[f'B{row}'].value
        if not prenom and not nom:
            continue
        if not any(is_green(sel[f'{col}{row}']) for col in ['E', 'F', 'G', 'H', 'I']):
            continue

        prenom_s = str(prenom).strip()
        if prenom_s:
            prenom_s = prenom_s[0].upper() + prenom_s[1:]
        nom_s    = str(nom).strip()
        full_nom = f'{prenom_s} {nom_s}'

        if full_nom.strip().lower() in existing:
            continue  # déjà traité

        age       = parse_age(sel[f'C{row}'].value)
        role      = str(sel[f'D{row}'].value or '').strip()
        selectors = [n for c, n in zip(['E', 'F', 'G', 'H', 'I'],
                                        ['Marilou', 'Phil', 'Charles', 'Isabelle', 'Amélie'])
                     if is_green(sel[f'{c}{row}'])]

        key       = (prenom_s.lower(), nom_s.lower())
        resp_data = resp_index.get(key, {})
        photo_url = resp_data.get('photo1')
        extra     = resp_data.get('extra', {})
        file_id   = None
        if photo_url:
            m = re.search(r'[?&]id=([a-zA-Z0-9_\-]+)', photo_url)
            if m:
                file_id = m.group(1)

        new_candidates.append({
            'nom':       full_nom,
            'age':       age,
            'role':      role,
            'selectors': selectors,
            'file_id':   file_id,
            'extra':     extra,
        })

    print(json.dumps(new_candidates, ensure_ascii=False, indent=2))


def cmd_process_photo(mcp_photo_file: str, nom: str) -> str:
    """Extrait, convertit si HEIC, uploade sur Supabase Storage. Retourne URL."""
    img = extract_img_from_mcp(mcp_photo_file)
    magic = img[:12]

    # Détection HEIC (ftyp box)
    is_heic = (len(magic) >= 8 and magic[4:8] == b'ftyp')
    is_jpg  = magic[:2] == b'\xff\xd8'
    is_png  = magic[:4] == b'\x89PNG'

    if is_heic:
        with tempfile.NamedTemporaryFile(suffix='.heic', delete=False) as tmp:
            tmp.write(img)
            heic_path = tmp.name
        jpg_path = heic_path.replace('.heic', '.jpg')
        result = subprocess.run(
            ['sips', '-s', 'format', 'jpeg', '-s', 'formatOptions', '90',
             heic_path, '--out', jpg_path],
            capture_output=True,
        )
        if result.returncode != 0:
            raise RuntimeError(f'sips: {result.stderr.decode()}')
        with open(jpg_path, 'rb') as f:
            img = f.read()
        os.unlink(heic_path)
        os.unlink(jpg_path)
        ext, mime = 'jpg', 'image/jpeg'
    elif is_jpg:
        ext, mime = 'jpg', 'image/jpeg'
    elif is_png:
        ext, mime = 'png', 'image/png'
    else:
        ext, mime = 'jpg', 'image/jpeg'

    import time, random, string
    safe = re.sub(r'[^a-z0-9]+', '_', nom.lower()).strip('_')
    rand = ''.join(random.choices(string.ascii_lowercase + string.digits, k=6))
    path = f'casting/{safe}_{rand}.{ext}'
    r = requests.post(
        f'{SUPABASE_URL}/storage/v1/object/{STORAGE_BUCKET}/{path}',
        headers={'apikey': STORAGE_KEY, 'Authorization': f'Bearer {STORAGE_KEY}',
                 'Content-Type': mime, 'x-upsert': 'true'},
        data=img,
        timeout=90,
    )
    r.raise_for_status()
    return f'{SUPABASE_URL}/storage/v1/object/public/{STORAGE_BUCKET}/{path}'


def roles_to_characters(role_str: str) -> list:
    """Extrait la liste de clés personnage à partir du texte du rôle.
    Ex: 'Premier rôle hockey - Louis, 16 ans, Deuxième rôle – Zack, 16 ans'
        → ['LOUIS', 'ZACK']
    """
    chars = []
    # Cherche les mots-clés de personnages précédés d'un tiret (- ou –)
    for word, key in _ROLE_CHAR_MAP.items():
        pattern = r'[-–]\s*' + re.escape(word)
        if re.search(pattern, role_str, re.I):
            if key not in chars:
                chars.append(key)
    return chars


def update_shortlists(nom: str, agence: str, role_str: str) -> list:
    """Ajoute le candidat dans les shortlists des personnages correspondant au rôle.
    Retourne la liste des personnages mis à jour."""
    characters = roles_to_characters(role_str)
    if not characters:
        return []

    r = requests.get(
        f'{SUPABASE_URL}/rest/v1/project_data',
        params={'key': f'eq.{PERSO_KEY}', 'select': 'data'},
        headers=SB_HEADERS, timeout=15,
    )
    r.raise_for_status()
    perso_data = r.json()[0]['data']

    updated = []
    for char_key in characters:
        if char_key not in perso_data:
            continue
        shortlist = perso_data[char_key].setdefault('shortlist', [])
        already = any(e.get('nom', '').strip().lower() == nom.strip().lower()
                      for e in shortlist)
        if not already:
            shortlist.append({'nom': nom, 'agence': agence})
            updated.append(char_key)

    if updated:
        now = datetime.now(timezone.utc).isoformat()
        r2 = requests.post(
            f'{SUPABASE_URL}/rest/v1/project_data',
            json={'key': PERSO_KEY, 'data': perso_data, 'updated_at': now},
            headers={**SB_HEADERS, 'Prefer': 'resolution=merge-duplicates'},
            timeout=30,
        )
        r2.raise_for_status()

    return updated


def cmd_save(nom: str, age_s: str, role: str, selectors_s: str, photo_url: str, extra_json: str = ''):
    """Ajoute un candidat dans Supabase. extra_json → fiche_cs sur le talent."""
    # Charger tous les talents
    r = requests.get(
        f'{SUPABASE_URL}/rest/v1/project_data',
        params={'key': f'eq.{CASTING_KEY}', 'select': 'data'},
        headers=SB_HEADERS, timeout=15,
    )
    r.raise_for_status()
    rows    = r.json()
    talents = rows[0]['data']['talents'] if rows else []
    max_id  = max((t.get('id', 0) for t in talents), default=0)

    # Vérifier doublon
    existing_lower = {t['nom'].strip().lower() for t in talents if t.get('agence') == 'Casting Sauvage'}
    if nom.strip().lower() in existing_lower:
        print(f'SKIP (déjà présent) : {nom}')
        return

    age       = int(age_s) if age_s and age_s.isdigit() else None
    selectors = [s.strip() for s in selectors_s.split(',') if s.strip()]
    prenom    = nom.split()[0] if nom else ''
    fiche_cs  = json.loads(extra_json) if extra_json.strip() else {}

    notes_parts = []
    if role:
        notes_parts.append(role)
    if selectors:
        notes_parts.append(f"Sélectionné·e par : {', '.join(selectors)}")

    talent = {
        'id':     max_id + 1,
        'nom':    nom,
        'agence': 'Casting Sauvage',
    }
    genre = guess_genre(prenom)
    if genre:
        talent['genre'] = genre
    if age:
        talent['age_min'] = age
        talent['age_max'] = age
    if photo_url:
        talent['photo_url'] = photo_url
    if notes_parts:
        talent['notes'] = ' | '.join(notes_parts)
    if fiche_cs:
        talent['fiche_cs'] = fiche_cs

    talents.append(talent)
    now = datetime.now(timezone.utc).isoformat()
    r2 = requests.post(
        f'{SUPABASE_URL}/rest/v1/project_data',
        json={'key': CASTING_KEY, 'data': {'talents': talents}, 'updated_at': now},
        headers={**SB_HEADERS, 'Prefer': 'resolution=merge-duplicates'},
        timeout=30,
    )
    r2.raise_for_status()
    print(f'✓ Ajouté : {nom} (id {max_id + 1})')

    # Mise à jour des shortlists de personnages
    if role:
        updated_chars = update_shortlists(nom, 'Casting Sauvage', role)
        if updated_chars:
            print(f'  → Shortlists : {", ".join(updated_chars)}')


def cmd_sync_shortlists():
    """Rattache rétroactivement tous les candidats Casting Sauvage aux shortlists."""
    r = requests.get(
        f'{SUPABASE_URL}/rest/v1/project_data',
        params={'key': f'eq.{CASTING_KEY}', 'select': 'data'},
        headers=SB_HEADERS, timeout=15,
    )
    r.raise_for_status()
    talents = r.json()[0]['data'].get('talents', [])
    sauvage = [t for t in talents if t.get('agence') == 'Casting Sauvage']
    print(f'{len(sauvage)} candidats Casting Sauvage à vérifier…')

    total_updated = 0
    for t in sauvage:
        nom   = t.get('nom', '')
        notes = t.get('notes', '')
        # Le rôle est la première partie des notes (avant " | Sélectionné·e par")
        role = notes.split(' | ')[0] if notes else ''
        if not role:
            continue
        chars = roles_to_characters(role)
        if not chars:
            continue
        updated = update_shortlists(nom, 'Casting Sauvage', role)
        if updated:
            print(f'  {nom} → {", ".join(updated)}')
            total_updated += 1

    print(f'\n✓ {total_updated} candidats ajoutés aux shortlists.')


def cmd_enrich(mcp_xlsx_file: str):
    """Enrichit rétroactivement les candidats CS sans fiche_cs depuis le XLSX."""
    try:
        import openpyxl
    except ImportError:
        sys.exit('openpyxl requis : pip3 install openpyxl')

    xlsx_bytes = extract_xlsx_from_mcp(mcp_xlsx_file)
    xlsx_path  = '/tmp/_casting_enrich_tmp.xlsx'
    with open(xlsx_path, 'wb') as f:
        f.write(xlsx_bytes)

    wb  = openpyxl.load_workbook(xlsx_path, data_only=True)
    rep = wb['Réponses au formulaire 1']

    # Index réponses — dernier envoi gagne
    resp_index = {}
    for row in range(2, rep.max_row + 1):
        prenom = rep.cell(row, 2).value
        nom    = rep.cell(row, 3).value
        if not prenom and not nom:
            continue
        key = (str(prenom).strip().lower(), str(nom).strip().lower())
        resp_index[key] = _parse_extra_from_response_row(rep, row)

    # Charger les talents existants
    r = requests.get(
        f'{SUPABASE_URL}/rest/v1/project_data',
        params={'key': f'eq.{CASTING_KEY}', 'select': 'data'},
        headers=SB_HEADERS, timeout=15,
    )
    r.raise_for_status()
    rows    = r.json()
    talents = rows[0]['data'].get('talents', []) if rows else []

    updated = 0
    for t in talents:
        if t.get('agence') != 'Casting Sauvage':
            continue
        if t.get('fiche_cs'):
            continue
        nom_full = t.get('nom', '').strip()
        parts = nom_full.split()
        if len(parts) < 2:
            continue
        key   = (parts[0].lower(), ' '.join(parts[1:]).lower())
        extra = resp_index.get(key)
        if extra:
            t['fiche_cs'] = extra
            print(f'  ✓ {nom_full}')
            updated += 1
        else:
            print(f'  ? {nom_full} — non trouvé dans Réponses')

    if updated:
        now = datetime.now(timezone.utc).isoformat()
        r2 = requests.post(
            f'{SUPABASE_URL}/rest/v1/project_data',
            json={'key': CASTING_KEY, 'data': {'talents': talents}, 'updated_at': now},
            headers={**SB_HEADERS, 'Prefer': 'resolution=merge-duplicates'},
            timeout=30,
        )
        r2.raise_for_status()
        print(f'\n✓ {updated} candidats enrichis.')
    else:
        print('\nAucun candidat à enrichir.')


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    mode = sys.argv[1]

    if mode == 'check':
        if len(sys.argv) < 3:
            sys.exit('Usage: casting_watcher.py check <mcp_xlsx_file>')
        cmd_check(sys.argv[2])

    elif mode == 'process-photo':
        if len(sys.argv) < 4:
            sys.exit('Usage: casting_watcher.py process-photo <mcp_photo_file> <nom>')
        url = cmd_process_photo(sys.argv[2], sys.argv[3])
        print(url)

    elif mode == 'save':
        if len(sys.argv) < 7:
            sys.exit('Usage: casting_watcher.py save <nom> <age> <role> <selectors> <photo_url> [extra_json]')
        extra_json = sys.argv[7] if len(sys.argv) > 7 else ''
        cmd_save(sys.argv[2], sys.argv[3], sys.argv[4], sys.argv[5], sys.argv[6], extra_json)

    elif mode == 'enrich':
        if len(sys.argv) < 3:
            sys.exit('Usage: casting_watcher.py enrich <mcp_xlsx_file>')
        cmd_enrich(sys.argv[2])

    elif mode == 'sync-shortlists':
        cmd_sync_shortlists()

    else:
        print(__doc__)
        sys.exit(1)
